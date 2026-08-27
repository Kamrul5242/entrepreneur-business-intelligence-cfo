#!/usr/bin/env python3
"""Build the premium multi-currency CFO dashboard workbook.

KNOWN ISSUE: under openpyxl 3.1.5 the workbook this writes cannot be opened by
Excel 16.0.20326 - not even in repair mode. Reproduced with the unmodified
original version of this script, so it is not caused by the Trend or Start Here
sheets. ALWAYS open the output in Excel before shipping it. The committed
workbook is the original build, which opens correctly; do not overwrite it with
an unverified rebuild."""
import argparse
import os
import sys

_DEFAULT_OUT = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "assets", "cfo-premium-dashboard.xlsx"))

_ap = argparse.ArgumentParser(
    description="Rebuild the CFO premium dashboard workbook.",
    epilog="The default target is a released, SHA-256 signed asset. Rebuilding "
           "it changes its hash and makes verify_signature.py report TAMPERED, "
           "so overwriting requires --force.")
_ap.add_argument("--output", "-o", default=_DEFAULT_OUT,
                 help="where to write the workbook (default: %(default)s)")
_ap.add_argument("--force", "-f", action="store_true",
                 help="overwrite the output file if it already exists")
ARGS = _ap.parse_args()

if os.path.exists(ARGS.output) and not ARGS.force:
    sys.exit(
        "refusing to overwrite an existing workbook:\n"
        "  {}\n"
        "This file is recorded in SIGNATURE.json; rebuilding changes its hash.\n"
        "Re-run with --force to overwrite, or -o PATH to write elsewhere."
        .format(ARGS.output))
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             NamedStyle, Protection)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.comments import Comment

SIG = "MKH-EBIC-2.2.0"
AUTHOR = "Md Kamrul Hasan"
GH = "https://github.com/Kamrul5242"

# ---- palette: deep navy / slate / gold accent ----
NAVY   = "0F2438"
NAVY2  = "1B3A57"
SLATE  = "2E4A63"
GOLD   = "C9A227"
LIGHT  = "F4F6F8"
BAND   = "EAEEF2"
WHITE  = "FFFFFF"
GREEN  = "1E7B45"
RED    = "B3261E"
AMBER  = "B26A00"
BORDER = "C7D0D9"

F = "Arial"
thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(ws, ref, *, bold=False, size=10, color="1A1A1A", fill=None,
          align=None, fmt=None, border=True, wrap=False, italic=False):
    for row in ws[ref]:
        for c in row:
            c.font = Font(name=F, bold=bold, size=size, color=color, italic=italic)
            if fill: c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(horizontal=align or "left", vertical="center", wrap_text=wrap)
            if fmt: c.number_format = fmt
            if border: c.border = box

def banner(ws, row, text, span=8, sub=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=F, bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30
    if sub:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=span)
        s = ws.cell(row=row+1, column=1, value=sub)
        s.font = Font(name=F, size=9, italic=True, color=WHITE)
        s.fill = PatternFill("solid", fgColor=NAVY2)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row+1].height = 18

def section(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=F, bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def sigfooter(ws, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1,
                value=f"◆ {AUTHOR}  ·  {GH}  ·  Digital Signature {SIG}  ·  "
                      f"Attribution required under MIT License")
    c.font = Font(name=F, size=8, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.protection = Protection(locked=True)
    ws.row_dimensions[row].height = 20

wb = Workbook()

# =====================================================================
# SHEET 1 — SETUP (currency + inputs)
# =====================================================================
s = wb.active
s.title = "1. Setup"
s.sheet_view.showGridLines = False
widths = {"A":34,"B":18,"C":16,"D":16,"E":16,"F":16,"G":16,"H":22}
for k,v in widths.items(): s.column_dimensions[k].width = v

banner(s, 1, "CFO DASHBOARD  ·  SETUP & INPUTS",
       sub="Fill ONLY the yellow cells. Everything else calculates automatically.")

section(s, 4, "CURRENCY")
s["A5"] = "Select currency"
s["B5"] = "BDT"
s["A6"] = "Symbol (auto)"
s["B6"] = '=IFERROR(INDEX(Ref!$B$2:$B$35,MATCH($B$5,Ref!$A$2:$A$35,0)),"?")'
# Currency symbols span many scripts (Bengali, Arabic, Devanagari, CJK, Won,
# Dong...). Arial does not carry all of them on every renderer. Give this
# cell a font family list with broad Unicode coverage so BDT/AED/INR/KRW/etc
# all render correctly everywhere the file is opened, not just where a
# system-level font substitution happens to kick in.
# Excel font names must be a SINGLE family, max 31 chars, no commas. A
# CSS-style stack here produced <name val="Noto Sans,FreeSans,..."/>, which
# Excel rejects outright - the workbook would not open at all. Excel does
# its own font substitution when a family is missing, so one name is enough.
SYMFONT = "Nirmala UI"
s["A7"] = "Currency name (auto)"
s["B7"] = '=IFERROR(INDEX(Ref!$C$2:$C$35,MATCH($B$5,Ref!$A$2:$A$35,0)),"Unknown")'
s["A8"] = "Reporting period"
s["B8"] = "2026-03 (monthly)"

dv = DataValidation(type="list", formula1="=Ref!$A$2:$A$35", allow_blank=False,
                    showDropDown=False)
dv.prompt = "Pick any of 34 currencies"; dv.promptTitle = "Currency"
s.add_data_validation(dv); dv.add(s["B5"])

style(s, "A5:A8", bold=True, fill=LIGHT)
style(s, "B5:B5", bold=True, color="0000FF", fill="FFFF00", align="center")
style(s, "B8:B8", color="0000FF", fill="FFFF00", align="center")
style(s, "B6:B7", align="center", fill=BAND)
s["B6"].font = Font(name=SYMFONT, bold=True, size=10, color="1A1A1A")
s["B5"].comment = Comment(f"Change this to switch every currency symbol in the "
                          f"workbook.\n{AUTHOR} · {SIG}", AUTHOR)

section(s, 10, "INCOME  ·  enter your numbers")
rows_income = [
    ("Gross Revenue", 920000, "All sales before returns"),
    ("Returns & Refunds", 70000, "Enter positive"),
    ("Discounts Given", 0, ""),
]
r = 11
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Amount")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
for name, val, note in rows_income:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
INC_START = 12
s.cell(row=r, column=1, value="NET REVENUE")
s.cell(row=r, column=2, value=f"=B{INC_START}-B{INC_START+1}-B{INC_START+2}")
NET_REV = r
style(s, f"A{r}:C{r}", bold=True, fill=BAND)

section(s, r+2, "COST OF GOODS SOLD")
r += 3
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Amount")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
COGS_START = r
for name, val, note in [("Product / Material (LANDED)", 380000, "incl. freight, duty, clearing"),
                        ("Direct Labour", 40000, "Production staff only"),
                        ("Packaging", 22000, "")]:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
COGS_END = r-1
s.cell(row=r, column=1, value="TOTAL COGS")
s.cell(row=r, column=2, value=f"=SUM(B{COGS_START}:B{COGS_END})")
COGS_TOT = r
style(s, f"A{r}:C{r}", bold=True, fill=BAND)

section(s, r+2, "OTHER VARIABLE COSTS  ·  the ones founders forget")
r += 3
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Amount")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
VAR_START = r
for name, val, note in [("Outbound Shipping", 102000, "Your share"),
                        ("Payment Gateway Fees", 21250, "% + fixed"),
                        ("Marketplace Commission", 0, ""),
                        ("RTO / Failed Delivery", 0, "Two-way shipping loss")]:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
VAR_END = r-1
s.cell(row=r, column=1, value="TOTAL OTHER VARIABLE")
s.cell(row=r, column=2, value=f"=SUM(B{VAR_START}:B{VAR_END})")
VAR_TOT = r
style(s, f"A{r}:C{r}", bold=True, fill=BAND)

section(s, r+2, "OPERATING EXPENSES  ·  fixed")
r += 3
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Amount")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
OP_START = r
for name, val, note in [("Advertising Spend", 240000, "All platforms"),
                        ("Salaries (non-production)", 95000, ""),
                        ("Owner Compensation", 40000, "Market rate even if unpaid"),
                        ("Rent", 25000, ""),
                        ("Software & Subscriptions", 8000, ""),
                        ("Other Operating", 12000, ""),
                        ("Depreciation & Amortization", 0, "Non-cash")]:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
OP_END = r-1
ADS_ROW = OP_START
DA_ROW = OP_END
s.cell(row=r, column=1, value="TOTAL OPEX")
s.cell(row=r, column=2, value=f"=SUM(B{OP_START}:B{OP_END})")
OP_TOT = r
style(s, f"A{r}:C{r}", bold=True, fill=BAND)

section(s, r+2, "BELOW THE LINE  ·  and cash-only items")
r += 3
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Amount")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
BL_START = r
for name, val, note in [("Interest Expense", 12000, "P&L item"),
                        ("Tax", 0, ""),
                        ("Loan Principal Repaid", 35000, "CASH ONLY — not a P&L expense"),
                        ("Capital Expenditure", 0, "Cash only"),
                        ("Owner Drawings", 60000, "CASH ONLY — not a P&L expense")]:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
INT_ROW, TAX_ROW, PRIN_ROW, CAPEX_ROW, DRAW_ROW = BL_START, BL_START+1, BL_START+2, BL_START+3, BL_START+4

section(s, r+1, "OPERATIONS & BALANCE  ·  for unit economics and cash cycle")
r += 2
s.cell(row=r, column=1, value="Line item"); s.cell(row=r, column=2, value="Value")
s.cell(row=r, column=3, value="Note")
style(s, f"A{r}:C{r}", bold=True, fill=SLATE, color=WHITE, align="center")
r += 1
OPS_START = r
for name, val, note in [("Number of Orders", 1700, "count"),
                        ("New Customers", 1200, "count"),
                        ("Repeat Customers", 500, "count"),
                        ("Customers Lost (churn)", 90, "count"),
                        ("Customers at Period Start", 1500, "count"),
                        ("Cash & Bank (closing)", 850000, ""),
                        ("Accounts Receivable (avg)", 310000, ""),
                        ("Inventory at Cost (avg)", 640000, ""),
                        ("Accounts Payable (avg)", 220000, ""),
                        ("Total Debt", 700000, ""),
                        ("Current Assets", 1800000, ""),
                        ("Current Liabilities", 950000, ""),
                        ("Total Equity", 1400000, ""),
                        ("Days in Period", 30, "30 / 90 / 365")]:
    s.cell(row=r, column=1, value=name); s.cell(row=r, column=2, value=val)
    s.cell(row=r, column=3, value=note); r += 1
(ORD, NEWC, REPC, LOSTC, STARTC, CASH, AR, INV, AP, DEBT,
 CA, CL, EQ, DAYS) = [OPS_START + i for i in range(14)]

# input formatting: yellow/blue for all editable amount cells
for a, b in [(INC_START, INC_START+2), (COGS_START, COGS_END),
             (VAR_START, VAR_END), (OP_START, OP_END),
             (BL_START, BL_START+4), (OPS_START, OPS_START+13)]:
    style(s, f"A{a}:A{b}", fill=LIGHT)
    style(s, f"B{a}:B{b}", color="0000FF", fill="FFFF00", align="right",
          fmt='#,##0;(#,##0);-')
    style(s, f"C{a}:C{b}", size=9, italic=True, color="5A6672", wrap=True)

for tot in (NET_REV, COGS_TOT, VAR_TOT, OP_TOT):
    s.cell(row=tot, column=2).number_format = '#,##0;(#,##0);-'
    s.cell(row=tot, column=2).alignment = Alignment(horizontal="right")

s.cell(row=DAYS, column=2).number_format = '0'
for rr in (ORD, NEWC, REPC, LOSTC, STARTC):
    s.cell(row=rr, column=2).number_format = '#,##0'

LEG = r+1
s.merge_cells(start_row=LEG, start_column=1, end_row=LEG, end_column=3)
s.cell(row=LEG, column=1,
       value="LEGEND — Yellow fill + blue text = your input. "
             "Grey = calculated, do not edit. All figures in the currency chosen at B5.")
style(s, f"A{LEG}:C{LEG}", bold=True, size=9, fill=BAND, wrap=True)
s.row_dimensions[LEG].height = 28
sigfooter(s, LEG+2, span=3)
s.freeze_panes = "A4"

R = {"NET_REV":NET_REV,"COGS":COGS_TOT,"VAR":VAR_TOT,"OP":OP_TOT,"ADS":ADS_ROW,
     "DA":DA_ROW,"INT":INT_ROW,"TAX":TAX_ROW,"PRIN":PRIN_ROW,"CAPEX":CAPEX_ROW,
     "DRAW":DRAW_ROW,"ORD":ORD,"NEWC":NEWC,"REPC":REPC,"LOSTC":LOSTC,
     "STARTC":STARTC,"CASH":CASH,"AR":AR,"INV":INV,"AP":AP,"DEBT":DEBT,
     "CA":CA,"CL":CL,"EQ":EQ,"DAYS":DAYS}
S = lambda k: f"'1. Setup'!$B${R[k]}"
SYM = "'1. Setup'!$B$6"

# =====================================================================
# SHEET 2 — DASHBOARD
# =====================================================================
d = wb.create_sheet("2. Dashboard")
d.sheet_view.showGridLines = False
for k,v in {"A":30,"B":20,"C":14,"D":26,"E":20,"F":14,"G":24,"H":16}.items():
    d.column_dimensions[k].width = v

banner(d, 1, "EXECUTIVE DASHBOARD",
       sub="Live from Setup. Change any input and every figure here updates.")
d["A3"] = f'=""&{SYM}&"  ·  Period: "&\'1. Setup\'!$B$8'
style(d, "A3:H3", bold=True, size=10, fill=BAND, border=False)
d["A3"].font = Font(name=SYMFONT,
                    bold=True, size=10, color="1A1A1A")

# KPI cards
section(d, 5, "KEY NUMBERS", span=3)
# Money-valued rows first (0..5) so the waterfall chart below can reference a
# clean contiguous block. Percentage rows (Gross Margin, Net Margin) go last —
# plotting a 0-1 ratio on the same axis as six-figure currency amounts makes
# the percentage bar visually vanish, so they are deliberately excluded from
# the chart range rather than reordered around it.
kpis = [
    ("Net Revenue",      f"={S('NET_REV')}",                                  'money'),
    ("Gross Profit",     f"={S('NET_REV')}-{S('COGS')}",                      'money'),
    ("Contribution (pre-ad)", f"={S('NET_REV')}-{S('COGS')}-{S('VAR')}",      'money'),
    ("CM after Ads",     f"={S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('ADS')}",'money'),
    ("Operating Profit", f"={S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('OP')}", 'money'),
    ("Net Profit",       f"={S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('OP')}-{S('INT')}-{S('TAX')}", 'money'),
    ("Gross Margin",     f"=IFERROR(({S('NET_REV')}-{S('COGS')})/{S('NET_REV')},0)", 'pct'),
    ("Net Margin",       f"=IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('OP')}-{S('INT')}-{S('TAX')})/{S('NET_REV')},0)", 'pct'),
]
KPI_MONEY_ROWS = 6  # first N rows above are money-valued; chart uses only these
row = 6
d.cell(row=row, column=1, value="Metric"); d.cell(row=row, column=2, value="Value")
d.cell(row=row, column=3, value="Status")
style(d, f"A{row}:C{row}", bold=True, fill=SLATE, color=WHITE, align="center")
row += 1
KPI_START = row
for label, formula, kind in kpis:
    d.cell(row=row, column=1, value=label)
    d.cell(row=row, column=2, value=formula)
    d.cell(row=row, column=3, value=f'=IF(B{row}>0,"OK",IF(B{row}=0,"FLAT","LOSS"))')
    c = d.cell(row=row, column=2)
    c.number_format = ('0.0%' if kind == 'pct' else '#,##0;(#,##0);-')
    row += 1
KPI_END = row-1
style(d, f"A{KPI_START}:A{KPI_END}", bold=True, fill=LIGHT)
style(d, f"B{KPI_START}:B{KPI_END}", bold=True, align="right", size=11)
style(d, f"C{KPI_START}:C{KPI_END}", align="center", bold=True)
for rr in range(KPI_START, KPI_END+1):
    d.cell(row=rr, column=2).number_format = ('0.0%' if kpis[rr-KPI_START][2]=='pct'
                                              else '#,##0;(#,##0);-')
d.conditional_formatting.add(f"C{KPI_START}:C{KPI_END}",
    CellIsRule(operator="equal", formula=['"LOSS"'],
               font=Font(name=F, bold=True, color=WHITE),
               fill=PatternFill(bgColor=RED)))
d.conditional_formatting.add(f"C{KPI_START}:C{KPI_END}",
    CellIsRule(operator="equal", formula=['"OK"'],
               font=Font(name=F, bold=True, color=WHITE),
               fill=PatternFill(bgColor=GREEN)))

# Unit economics block (cols E-H)
d.merge_cells("E5:H5")
d["E5"] = "UNIT ECONOMICS"
d["E5"].font = Font(name=F, bold=True, size=11, color=WHITE)
d["E5"].fill = PatternFill("solid", fgColor=SLATE)
d["E5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

d["E6"] = "Metric"; d["F6"] = "Value"
style(d, "E6:F6", bold=True, fill=SLATE, color=WHITE, align="center")
ue = [
    ("AOV",                    f"=IFERROR({S('NET_REV')}/{S('ORD')},0)", '#,##0.00'),
    ("Blended CAC",            f"=IFERROR({S('ADS')}/{S('NEWC')},0)", '#,##0.00'),
    ("CM per order",           f"=IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('ADS')})/{S('ORD')},0)", '#,##0.00'),
    ("CM ratio (pre-ad)",      f"=IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')})/{S('NET_REV')},0)", '0.0%'),
    ("Break-even ROAS",        f"=IFERROR({S('NET_REV')}/({S('NET_REV')}-{S('COGS')}-{S('VAR')}),0)", '0.00"x"'),
    ("Actual ROAS (blended)",  f"=IFERROR({S('NET_REV')}/{S('ADS')},0)", '0.00"x"'),
    ("ROAS headroom",          f"=IFERROR({S('NET_REV')}/{S('ADS')}-{S('NET_REV')}/({S('NET_REV')}-{S('COGS')}-{S('VAR')}),0)", '0.00"x"'),
    ("Repeat rate",            f"=IFERROR({S('REPC')}/({S('NEWC')}+{S('REPC')}),0)", '0.0%'),
    ("Churn rate",             f"=IFERROR({S('LOSTC')}/{S('STARTC')},0)", '0.0%'),
]
rr = 7
for label, formula, fmt in ue:
    d.cell(row=rr, column=5, value=label)
    c = d.cell(row=rr, column=6, value=formula); c.number_format = fmt
    rr += 1
UE_END = rr-1
style(d, f"E7:E{UE_END}", fill=LIGHT)
style(d, f"F7:F{UE_END}", bold=True, align="right")
for i,(label,formula,fmt) in enumerate(ue):
    d.cell(row=7+i, column=6).number_format = fmt

# Cash & health
section(d, KPI_END+2, "CASH  ·  RUNWAY", span=3)
d.merge_cells(start_row=KPI_END+2, start_column=5, end_row=KPI_END+2, end_column=6)
_c = d.cell(row=KPI_END+2, column=5, value="LIQUIDITY & CASH CYCLE")
_c.font = Font(name=F, bold=True, size=11, color=WHITE)
_c.fill = PatternFill("solid", fgColor=SLATE)
_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
hr = KPI_END+3
d.cell(row=hr, column=1, value="Metric"); d.cell(row=hr, column=2, value="Value")
d.cell(row=hr, column=3, value="Status")
d.cell(row=hr, column=5, value="Metric"); d.cell(row=hr, column=6, value="Value")
style(d, f"A{hr}:C{hr}", bold=True, fill=SLATE, color=WHITE, align="center")
style(d, f"E{hr}:F{hr}", bold=True, fill=SLATE, color=WHITE, align="center")

NP = f"({S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('OP')}-{S('INT')}-{S('TAX')})"
cash_left = [
    ("Operating Cash Flow (approx)", f"={NP}+{S('DA')}", '#,##0;(#,##0);-'),
    ("Cash out below the line",      f"={S('PRIN')}+{S('CAPEX')}+{S('DRAW')}", '#,##0;(#,##0);-'),
    ("Net Cash Change",              f"={NP}+{S('DA')}-{S('PRIN')}-{S('CAPEX')}-{S('DRAW')}", '#,##0;(#,##0);-'),
    ("Monthly Net Burn",             f"=MAX(0,-({NP}+{S('DA')}-{S('PRIN')}-{S('CAPEX')}-{S('DRAW')}))", '#,##0;(#,##0);-'),
    ("Runway (months)",              f'=IF(MAX(0,-({NP}+{S("DA")}-{S("PRIN")}-{S("CAPEX")}-{S("DRAW")}))=0,"cash positive",{S("CASH")}/MAX(1,-({NP}+{S("DA")}-{S("PRIN")}-{S("CAPEX")}-{S("DRAW")})))', '0.0'),
]
rr = hr+1
CASH_START = rr
for label, formula, fmt in cash_left:
    d.cell(row=rr, column=1, value=label)
    c = d.cell(row=rr, column=2, value=formula); c.number_format = fmt
    d.cell(row=rr, column=3, value=f'=IF(ISTEXT(B{rr}),"OK",IF(B{rr}>=0,"OK","LOSS"))')
    rr += 1
CASH_END = rr-1
style(d, f"A{CASH_START}:A{CASH_END}", bold=True, fill=LIGHT)
style(d, f"B{CASH_START}:B{CASH_END}", bold=True, align="right")
style(d, f"C{CASH_START}:C{CASH_END}", align="center", bold=True)
for i,(l,f_,fmt) in enumerate(cash_left):
    d.cell(row=CASH_START+i, column=2).number_format = fmt
d.conditional_formatting.add(f"C{CASH_START}:C{CASH_END}",
    CellIsRule(operator="equal", formula=['"LOSS"'],
               font=Font(name=F, bold=True, color=WHITE),
               fill=PatternFill(bgColor=RED)))
d.conditional_formatting.add(f"C{CASH_START}:C{CASH_END}",
    CellIsRule(operator="equal", formula=['"OK"'],
               font=Font(name=F, bold=True, color=WHITE),
               fill=PatternFill(bgColor=GREEN)))

liq = [
    ("Current Ratio",   f"=IFERROR({S('CA')}/{S('CL')},0)", '0.00'),
    ("Quick Ratio",     f"=IFERROR(({S('CA')}-{S('INV')})/{S('CL')},0)", '0.00'),
    ("Debt / Equity",   f"=IFERROR({S('DEBT')}/{S('EQ')},0)", '0.00'),
    ("Interest Coverage", f"=IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')}-{S('OP')})/{S('INT')},0)", '0.00"x"'),
    ("DIO (days)",      f"=IFERROR({S('INV')}/{S('COGS')}*{S('DAYS')},0)", '0.0'),
    ("DSO (days)",      f"=IFERROR({S('AR')}/{S('NET_REV')}*{S('DAYS')},0)", '0.0'),
    ("DPO (days)",      f"=IFERROR({S('AP')}/{S('COGS')}*{S('DAYS')},0)", '0.0'),
    ("Cash Conversion Cycle", f"=IFERROR({S('INV')}/{S('COGS')}*{S('DAYS')}+{S('AR')}/{S('NET_REV')}*{S('DAYS')}-{S('AP')}/{S('COGS')}*{S('DAYS')},0)", '0.0'),
]
rr = hr+1
for label, formula, fmt in liq:
    d.cell(row=rr, column=5, value=label)
    c = d.cell(row=rr, column=6, value=formula); c.number_format = fmt
    rr += 1
LIQ_END = rr-1
style(d, f"E{hr+1}:E{LIQ_END}", fill=LIGHT)
style(d, f"F{hr+1}:F{LIQ_END}", bold=True, align="right")
for i,(l,f_,fmt) in enumerate(liq):
    d.cell(row=hr+1+i, column=6).number_format = fmt

# Break-even
BE = max(CASH_END, LIQ_END)+2
section(d, BE, "BREAK-EVEN")
d.cell(row=BE+1, column=1, value="Fixed costs (OPEX excl. ads)")
d.cell(row=BE+1, column=2, value=f"={S('OP')}-{S('ADS')}")
d.cell(row=BE+2, column=1, value="CM ratio (pre-ad)")
d.cell(row=BE+2, column=2, value=f"=IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')})/{S('NET_REV')},0)")
d.cell(row=BE+3, column=1, value="Break-even revenue")
d.cell(row=BE+3, column=2, value=f"=IFERROR(B{BE+1}/B{BE+2},0)")
d.cell(row=BE+4, column=1, value="Break-even orders")
d.cell(row=BE+4, column=2, value=f"=IFERROR(B{BE+3}/IFERROR({S('NET_REV')}/{S('ORD')},1),0)")
d.cell(row=BE+5, column=1, value="Margin of safety")
d.cell(row=BE+5, column=2, value=f"=IFERROR(({S('NET_REV')}-B{BE+3})/{S('NET_REV')},0)")
style(d, f"A{BE+1}:A{BE+5}", bold=True, fill=LIGHT)
style(d, f"B{BE+1}:B{BE+5}", bold=True, align="right")
for i, fmt in enumerate(['#,##0;(#,##0);-','0.0%','#,##0;(#,##0);-','#,##0','0.0%']):
    d.cell(row=BE+1+i, column=2).number_format = fmt

# Verdict
V = BE+7
section(d, V, "VERDICT")
# NOTE ON THE LOGIC — two facts that shape the branch order below:
#
# (1) "CM after ads < 0" and "Actual ROAS < Break-even ROAS" are algebraically
#     the SAME event (Break-even ROAS = NetRev/(NetRev-COGS-Var); the ROAS
#     inequality reduces exactly to CM_after_ads<0 for positive pre-ad
#     contribution). One branch covers both framings; a second branch testing
#     the ROAS form separately would be unreachable dead code.
#
# (2) NetProfit = CM_after_ads − FixedOpexExclAds − Interest − Tax, and those
#     three subtracted terms are never negative for a real business. So
#     CM_after_ads<0 is a STRICT SUBSET of NetProfit<0 — it can never be true
#     while NetProfit>=0. If NetProfit<0 were tested FIRST, it would shadow
#     the CM-after-ads branch completely and that branch could never fire
#     either. The fix is to test the MORE SPECIFIC condition first: that also
#     produces a more useful diagnosis, distinguishing "unit economics are
#     broken" (fix pricing/CAC/COGS) from "unit economics are fine but fixed
#     costs/interest/tax exceed them" (fix overhead, not the product).
d.merge_cells(start_row=V+1, start_column=1, end_row=V+1, end_column=8)
d.cell(row=V+1, column=1, value=(
    f'=IF(B{KPI_START+3}<0,"BROKEN UNIT ECONOMICS — contribution after ads is negative and actual ROAS is below break-even ROAS. Every order loses money before fixed costs are even counted. Stop scaling; fix pricing, CAC, or COGS first.",'
    f'IF(B{KPI_START+5}<0,"NET LOSS FROM OVERHEAD — each order is profitable after ads, but fixed costs, interest, or tax exceed that contribution. Fix overhead, not the product.",'
    f'"PROFITABLE — protect margin and scale toward the next constraint."))'))
style(d, f"A{V+1}:H{V+1}", bold=True, size=12, fill=BAND, wrap=True, align="center")
d.row_dimensions[V+1].height = 34

# chart
ch = BarChart(); ch.type="col"; ch.style=10
ch.title = "Profit Waterfall"
ch.y_axis.title = "Amount"; ch.x_axis.title = None
data = Reference(d, min_col=2, min_row=KPI_START, max_row=KPI_START+KPI_MONEY_ROWS-1)
cats = Reference(d, min_col=1, min_row=KPI_START, max_row=KPI_START+KPI_MONEY_ROWS-1)
ch.add_data(data, titles_from_data=False); ch.set_categories(cats)
ch.height = 8; ch.width = 20; ch.legend = None
d.add_chart(ch, f"A{V+3}")

sigfooter(d, V+22, span=8)
d.freeze_panes = "A5"

# =====================================================================
# SHEET 3 — SCENARIOS
# =====================================================================
sc = wb.create_sheet("3. Scenarios")
sc.sheet_view.showGridLines = False
for k,v in {"A":30,"B":22,"C":18,"D":18,"E":18,"F":18}.items():
    sc.column_dimensions[k].width = v
banner(sc, 1, "SCENARIO MODEL", span=6,
       sub="Edit the yellow drivers. Conservative / Base / Aggressive recalculate live.")

section(sc, 4, "DRIVERS", span=6)
sc["A5"]="Driver"; sc["B5"]="Conservative"; sc["C5"]="Base"; sc["D5"]="Aggressive"
style(sc,"A5:D5",bold=True,fill=SLATE,color=WHITE,align="center")
drv=[("Revenue change",-0.25,0.0,0.25),("COGS % change",0.04,0.0,-0.02),
     ("Ad spend change",0.10,0.0,0.20),("Fixed cost change",0.05,0.0,0.0)]
r=6
for n,a,b,c in drv:
    sc.cell(row=r,column=1,value=n); sc.cell(row=r,column=2,value=a)
    sc.cell(row=r,column=3,value=b); sc.cell(row=r,column=4,value=c); r+=1
DRV=6
style(sc,f"A{DRV}:A{r-1}",bold=True,fill=LIGHT)
style(sc,f"B{DRV}:D{r-1}",color="0000FF",fill="FFFF00",align="center",fmt='0.0%')

section(sc, r+1, "OUTCOME", span=6)
h=r+2
sc.cell(row=h,column=1,value="Line"); sc.cell(row=h,column=2,value="Conservative")
sc.cell(row=h,column=3,value="Base"); sc.cell(row=h,column=4,value="Aggressive")
style(sc,f"A{h}:D{h}",bold=True,fill=SLATE,color=WHITE,align="center")
o=h+1
lines=[("Net Revenue", lambda C: f"={S('NET_REV')}*(1+{C}{DRV})"),
       ("COGS",        lambda C: f"={S('COGS')}*(1+{C}{DRV})*(1+{C}{DRV+1})"),
       ("Other variable", lambda C: f"={S('VAR')}*(1+{C}{DRV})"),
       ("Ad spend",    lambda C: f"={S('ADS')}*(1+{C}{DRV+2})"),
       ("Fixed OPEX",  lambda C: f"=({S('OP')}-{S('ADS')})*(1+{C}{DRV+3})"),
       ]
for name, fn in lines:
    sc.cell(row=o,column=1,value=name)
    for i,C in enumerate("BCD"):
        sc.cell(row=o,column=2+i,value=fn(C))
    o+=1
NR_,CG_,VR_,AD_,FX_ = h+1,h+2,h+3,h+4,h+5
sc.cell(row=o,column=1,value="Gross Profit")
for i,C in enumerate("BCD"): sc.cell(row=o,column=2+i,value=f"={C}{NR_}-{C}{CG_}")
GP_=o; o+=1
sc.cell(row=o,column=1,value="Contribution after ads")
for i,C in enumerate("BCD"): sc.cell(row=o,column=2+i,value=f"={C}{NR_}-{C}{CG_}-{C}{VR_}-{C}{AD_}")
CM_=o; o+=1
sc.cell(row=o,column=1,value="Operating Profit")
for i,C in enumerate("BCD"): sc.cell(row=o,column=2+i,value=f"={C}{CM_}-{C}{FX_}")
OPP_=o; o+=1
sc.cell(row=o,column=1,value="Operating Margin")
for i,C in enumerate("BCD"): sc.cell(row=o,column=2+i,value=f"=IFERROR({C}{OPP_}/{C}{NR_},0)")
OM_=o; o+=1
sc.cell(row=o,column=1,value="Verdict")
for i,C in enumerate("BCD"):
    sc.cell(row=o,column=2+i,value=f'=IF({C}{OPP_}>0,"PROFIT","LOSS")')
VD_=o
style(sc,f"A{h+1}:A{VD_}",bold=True,fill=LIGHT)
style(sc,f"B{h+1}:D{VD_}",align="right",fmt='#,##0;(#,##0);-')
style(sc,f"B{OM_}:D{OM_}",align="right",fmt='0.0%')
style(sc,f"B{VD_}:D{VD_}",align="center",bold=True)
style(sc,f"B{GP_}:D{GP_}",bold=True,align="right",fmt='#,##0;(#,##0);-')
style(sc,f"B{OPP_}:D{OPP_}",bold=True,align="right",fmt='#,##0;(#,##0);-')
sc.conditional_formatting.add(f"B{VD_}:D{VD_}",
    CellIsRule(operator="equal",formula=['"LOSS"'],
               font=Font(name=F,bold=True,color=WHITE),
               fill=PatternFill(bgColor=RED)))
sc.conditional_formatting.add(f"B{VD_}:D{VD_}",
    CellIsRule(operator="equal",formula=['"PROFIT"'],
               font=Font(name=F,bold=True,color=WHITE),
               fill=PatternFill(bgColor=GREEN)))

section(sc, VD_+2, "PRICE SENSITIVITY  ·  how much volume can you afford to lose?", span=6)
p=VD_+3
sc.merge_cells(start_row=p, start_column=2, end_row=p, end_column=3)
sc.cell(row=p,column=1,value="Price increase")
sc.cell(row=p,column=2,value="Max volume loss tolerable")
style(sc,f"A{p}:C{p}",bold=True,fill=SLATE,color=WHITE,align="center")
cmratio=f"IFERROR(({S('NET_REV')}-{S('COGS')}-{S('VAR')})/{S('NET_REV')},0.3)"
for i,inc in enumerate([0.05,0.10,0.15,0.20]):
    sc.cell(row=p+1+i,column=1,value=inc)
    sc.merge_cells(start_row=p+1+i, start_column=2, end_row=p+1+i, end_column=3)
    sc.cell(row=p+1+i,column=2,value=f"=IFERROR({inc}/({cmratio}+{inc}),0)")
style(sc,f"A{p+1}:A{p+4}",bold=True,fill=LIGHT,align="center",fmt='0%')
style(sc,f"B{p+1}:C{p+4}",bold=True,align="center",fmt='0.0%')
sigfooter(sc, p+6, span=6)
sc.freeze_panes="A4"

# =====================================================================
# SHEET 4 — Ref (currency table)
# =====================================================================
# SHEET 4 — Trend (12 months)
# =====================================================================
tr = wb.create_sheet("4. Trend")
tr.sheet_view.showGridLines = False
banner(tr, 1, "12-MONTH TREND", span=13,
       sub="One column per month. Fill the yellow rows; the rest calculates. "
           "A single month cannot tell you whether a business is improving.")

MON_FIRST, MON_LAST = 2, 13          # columns B..M
TR_HDR = 4
tr.cell(row=TR_HDR, column=1, value="Line item")
for i in range(12):
    tr.cell(row=TR_HDR, column=MON_FIRST + i, value=f"M{i+1}")
style(tr, f"A{TR_HDR}:M{TR_HDR}", bold=True, fill=SLATE, color=WHITE, align="center")

TR_IN = TR_HDR + 1                    # rows 5..9 are inputs
TR_INPUTS = ["Net Revenue", "COGS", "Variable Costs (shipping, gateway, RTO)",
             "Ad Spend", "Fixed Operating Costs"]
for i, label in enumerate(TR_INPUTS):
    tr.cell(row=TR_IN + i, column=1, value=label)
style(tr, f"A{TR_IN}:A{TR_IN+4}", size=10)
style(tr, f"B{TR_IN}:M{TR_IN+4}", color="0000FF", fill="FFFF00", align="right",
      fmt='#,##0;(#,##0);-')

TR_CALC = TR_IN + 6                   # rows 11..15 computed
section(tr, TR_CALC - 1, "CALCULATED", span=13)
TR_ROWS = ["Contribution", "Contribution Margin %", "Operating Profit",
           "Operating Margin %", "Cumulative Operating Profit",
           "Revenue Growth vs Prior Month %"]
for i, label in enumerate(TR_ROWS):
    tr.cell(row=TR_CALC + i, column=1, value=label)

for i in range(12):
    col = get_column_letter(MON_FIRST + i)
    prev = get_column_letter(MON_FIRST + i - 1)
    rev, cogs_r, var_r = f"{col}{TR_IN}", f"{col}{TR_IN+1}", f"{col}{TR_IN+2}"
    ads_r, fix_r = f"{col}{TR_IN+3}", f"{col}{TR_IN+4}"
    contrib = f"{col}{TR_CALC}"
    tr[contrib] = f"={rev}-{cogs_r}-{var_r}"
    tr[f"{col}{TR_CALC+1}"] = f'=IFERROR({contrib}/{rev},"")'
    tr[f"{col}{TR_CALC+2}"] = f"={contrib}-{ads_r}-{fix_r}"
    tr[f"{col}{TR_CALC+3}"] = f'=IFERROR({col}{TR_CALC+2}/{rev},"")'
    if i == 0:
        tr[f"{col}{TR_CALC+4}"] = f"={col}{TR_CALC+2}"
        tr[f"{col}{TR_CALC+5}"] = ""
    else:
        tr[f"{col}{TR_CALC+4}"] = f"={prev}{TR_CALC+4}+{col}{TR_CALC+2}"
        tr[f"{col}{TR_CALC+5}"] = f'=IFERROR({rev}/{prev}{TR_IN}-1,"")'

style(tr, f"A{TR_CALC}:A{TR_CALC+5}", bold=True, size=10)
money = '#,##0;(#,##0);-'
for r in (TR_CALC, TR_CALC + 2, TR_CALC + 4):
    style(tr, f"B{r}:M{r}", align="right", fmt=money, fill=BAND)
for r in (TR_CALC + 1, TR_CALC + 3, TR_CALC + 5):
    style(tr, f"B{r}:M{r}", align="right", fmt='0.0%', fill=LIGHT)

# red when a month loses money
tr.conditional_formatting.add(
    f"B{TR_CALC+2}:M{TR_CALC+2}",
    CellIsRule(operator="lessThan", formula=["0"],
               font=Font(name=F, bold=True, color=RED)))

TR_CH = TR_CALC + 8
section(tr, TR_CH - 1, "REVENUE AND OPERATING PROFIT BY MONTH", span=13)
lc = LineChart(); lc.style = 12
lc.y_axis.title = "Amount"; lc.x_axis.title = "Month"
lc.height, lc.width = 8.5, 26
_data = Reference(tr, min_col=1, max_col=MON_LAST,
                  min_row=TR_IN, max_row=TR_IN)
_prof = Reference(tr, min_col=1, max_col=MON_LAST,
                  min_row=TR_CALC + 2, max_row=TR_CALC + 2)
lc.add_data(_data, titles_from_data=True, from_rows=True)
lc.add_data(_prof, titles_from_data=True, from_rows=True)
lc.set_categories(Reference(tr, min_col=MON_FIRST, max_col=MON_LAST,
                            min_row=TR_HDR, max_row=TR_HDR))
tr.add_chart(lc, f"A{TR_CH}")

tr.cell(row=TR_CH + 18, column=1,
        value="Read the direction, not the level. Three months of falling "
              "contribution margin is a structural problem; one month is noise.")
style(tr, f"A{TR_CH+18}:M{TR_CH+18}", size=9, italic=True, border=False)
sigfooter(tr, TR_CH + 20, span=13)
tr.column_dimensions["A"].width = 34
for i in range(12):
    tr.column_dimensions[get_column_letter(MON_FIRST + i)].width = 12
tr.freeze_panes = "B5"

# =====================================================================
ref = wb.create_sheet("Ref")
ref.sheet_view.showGridLines = False
ref["A1"]="Code"; ref["B1"]="Symbol"; ref["C1"]="Name"; ref["D1"]="Decimals"
CUR=[("BDT","৳","Bangladeshi Taka",2),("INR","₹","Indian Rupee",2),
("PKR","₨","Pakistani Rupee",2),("LKR","Rs","Sri Lankan Rupee",2),
("NPR","रू","Nepalese Rupee",2),("USD","$","US Dollar",2),
("EUR","€","Euro",2),("GBP","£","Pound Sterling",2),
("CAD","C$","Canadian Dollar",2),("AUD","A$","Australian Dollar",2),
("AED","د.إ","UAE Dirham",2),("SAR","SR","Saudi Riyal",2),
("QAR","QR","Qatari Riyal",2),("KWD","KD","Kuwaiti Dinar",3),
("BHD","BD","Bahraini Dinar",3),("OMR","OR","Omani Rial",3),
("MYR","RM","Malaysian Ringgit",2),("SGD","S$","Singapore Dollar",2),
("THB","฿","Thai Baht",2),("IDR","Rp","Indonesian Rupiah",0),
("VND","₫","Vietnamese Dong",0),("PHP","₱","Philippine Peso",2),
("CNY","CN¥","Chinese Yuan",2),("JPY","JP¥","Japanese Yen",0),
("KRW","₩","Korean Won",0),("TRY","₺","Turkish Lira",2),
("EGP","E£","Egyptian Pound",2),("NGN","₦","Nigerian Naira",2),
("ZAR","R","South African Rand",2),("KES","KSh","Kenyan Shilling",2),
("BRL","R$","Brazilian Real",2),("MXN","Mex$","Mexican Peso",2),
("CHF","CHF","Swiss Franc",2),("SEK","kr","Swedish Krona",2)]
for i,(c,s_,n,dd) in enumerate(CUR, start=2):
    ref.cell(row=i,column=1,value=c); ref.cell(row=i,column=2,value=s_)
    ref.cell(row=i,column=3,value=n); ref.cell(row=i,column=4,value=dd)
style(ref,"A1:D1",bold=True,fill=SLATE,color=WHITE,align="center")
style(ref,f"A2:D{len(CUR)+1}",size=9)
for k,v in {"A":10,"B":10,"C":26,"D":10}.items(): ref.column_dimensions[k].width=v

# =====================================================================
# SHEET 5 — Signature (protected)
# =====================================================================
sg = wb.create_sheet("Signature")
sg.sheet_view.showGridLines = False
for k,v in {"A":26,"B":62}.items(): sg.column_dimensions[k].width=v
banner(sg,1,"DIGITAL SIGNATURE  ·  DO NOT MODIFY",span=2)
meta=[("Author",AUTHOR),("GitHub",GH),("Skill","entrepreneur-business-intelligence-cfo"),
      ("Version","2.2.0"),("Signature ID",SIG),("License","MIT — attribution required"),
      ("Default currency","BDT (all 34 supported)"),
      ("Notice","This workbook and its formula design are the work of "
                "Md Kamrul Hasan. The MIT License permits reuse and modification "
                "provided this attribution is retained in all copies and "
                "substantial portions.")]
r=3
for k,v in meta:
    sg.cell(row=r,column=1,value=k); sg.cell(row=r,column=2,value=v); r+=1
style(sg,f"A3:A{r-1}",bold=True,fill=LIGHT)
style(sg,f"B3:B{r-1}",wrap=True)
sg.row_dimensions[r-1].height=64
sigfooter(sg,r+1,span=2)
for row in sg.iter_rows():
    for c in row: c.protection = Protection(locked=True)
sg.protection.sheet = True
sg.protection.password = "MKH-EBIC"
sg.protection.enable()

# =====================================================================
# SHEET 0 — Start Here  (guide; created last, moved to the front)
# =====================================================================
sh = wb.create_sheet("0. Start Here", 0)
sh.sheet_view.showGridLines = False
banner(sh, 1, "START HERE", span=8,
       sub="You do not need to understand accounting to use this. "
           "Fill eight numbers and read the verdict.")

section(sh, 4, "WHAT THIS IS", span=8)
sh["A5"] = ("A CFO in a spreadsheet. You type what your business earned and spent. "
            "It tells you whether you are making money, whether each order makes "
            "money, how long your cash lasts, and what to fix first.")
sh.merge_cells("A5:H5")
style(sh, "A5:H5", size=10, wrap=True, border=False)
sh.row_dimensions[5].height = 46

section(sh, 7, "THREE STEPS", span=8)
STEPS = [
    ("1", "Go to the sheet '1. Setup'. Type over the YELLOW cells only.",
     "Everything that is not yellow calculates itself. You cannot break it by "
     "typing in yellow."),
    ("2", "Start with the eight numbers marked START HERE in column C.",
     "The other rows are already filled with example numbers. Replace them as "
     "you find them; the dashboard works before you finish."),
    ("3", "Read '2. Dashboard' from the top.",
     "The coloured banner names your single biggest problem in one sentence."),
]
r = 8
for num, title, detail in STEPS:
    sh.cell(row=r, column=1, value=num)
    sh.cell(row=r, column=2, value=title)
    sh.cell(row=r + 1, column=2, value=detail)
    style(sh, f"A{r}:A{r}", bold=True, size=16, color=WHITE, fill=GOLD, align="center")
    style(sh, f"B{r}:H{r}", bold=True, size=11)
    style(sh, f"B{r+1}:H{r+1}", size=9, italic=True, color="4A5A6A", wrap=True)
    sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    sh.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=8)
    sh.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    sh.row_dimensions[r + 1].height = 26
    r += 2

section(sh, 15, "THE EIGHT NUMBERS YOU REALLY NEED", span=8)
sh.cell(row=16, column=1, value="On '1. Setup'")
sh.cell(row=16, column=2, value="What it means in plain words")
sh.cell(row=16, column=6, value="Where to find it")
style(sh, "A16:H16", bold=True, fill=SLATE, color=WHITE, size=9)
sh.merge_cells("B16:E16"); sh.merge_cells("F16:H16")
EIGHT = [
    ("Gross Revenue", "Everything customers paid you, before refunds.",
     "Your sales report or bank deposits"),
    ("Product / Material (LANDED)", "What the goods cost you delivered to your "
     "door - price plus freight plus duty, not the invoice alone.",
     "Supplier invoices plus shipping and customs"),
    ("Outbound Shipping", "Courier cost you actually pay, after what the "
     "customer covers.", "Courier bill"),
    ("Advertising Spend", "Every platform added together.",
     "Meta, Google, TikTok billing"),
    ("Salaries (non-production)", "People who are paid whether or not you sell "
     "anything.", "Payroll"),
    ("Rent", "Shop, office or warehouse.", "Rent receipt"),
    ("Number of Orders", "How many orders in the period. This turns totals into "
     "per-order economics.", "Order dashboard"),
    ("Cash & Bank (closing)", "What is actually in the bank on the last day.",
     "Bank statement"),
]
r = 17
for label, meaning, where in EIGHT:
    sh.cell(row=r, column=1, value=label)
    sh.cell(row=r, column=2, value=meaning)
    sh.cell(row=r, column=6, value=where)
    sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    sh.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    r += 1
style(sh, f"A17:A{r-1}", bold=True, size=9, fill=LIGHT)
style(sh, f"B17:E{r-1}", size=9, wrap=True)
style(sh, f"F17:H{r-1}", size=9, italic=True, color="4A5A6A", wrap=True)
for rr in range(17, r):
    sh.row_dimensions[rr].height = 26

section(sh, r + 1, "WHAT EACH SHEET DOES", span=8)
GUIDE = [
    ("1. Setup", "Where you type. Yellow cells only."),
    ("2. Dashboard", "The answer. Profit, unit economics, cash, runway, verdict."),
    ("3. Scenarios", "What if sales fall 25%? What if ads get cheaper?"),
    ("4. Trend", "Twelve months side by side. Is it getting better or worse?"),
    ("Ref", "The 34-currency table. You do not need to touch it."),
    ("Signature", "Who built this. Locked."),
]
rr = r + 2
for name, what in GUIDE:
    sh.cell(row=rr, column=1, value=name)
    sh.cell(row=rr, column=2, value=what)
    sh.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
    rr += 1
style(sh, f"A{r+2}:A{rr-1}", bold=True, size=9, fill=BAND)
style(sh, f"B{r+2}:H{rr-1}", size=9)

section(sh, rr + 1, "READ THIS BEFORE YOU TRUST THE NUMBER", span=8)
sh.cell(row=rr + 2, column=1,
        value="Profit is not cash. You can show a profit here and still have an "
              "empty bank account, because loan principal, owner drawings and "
              "stock purchases spend cash without appearing as costs. If the "
              "profit line and your bank balance disagree, believe the bank.")
style(sh, f"A{rr+2}:H{rr+2}", size=10, wrap=True, color=RED, bold=True)
sh.merge_cells(start_row=rr + 2, start_column=1, end_row=rr + 2, end_column=8)
sh.row_dimensions[rr + 2].height = 44

sh.cell(row=rr + 4, column=1,
        value="বাংলা:  শুধু হলুদ ঘরগুলোতে আপনার সংখ্যা লিখুন। বাকি সব নিজে হিসাব হয়ে যাবে। "
              "'২. Dashboard' শীটে গিয়ে উপরের রঙিন লাইনটা পড়ুন — ওখানে আপনার "
              "সবচেয়ে বড় সমস্যাটা এক লাইনে লেখা থাকবে।")
style(sh, f"A{rr+4}:H{rr+4}", size=10, wrap=True, fill=LIGHT)
sh.merge_cells(start_row=rr + 4, start_column=1, end_row=rr + 4, end_column=8)
sh.row_dimensions[rr + 4].height = 34

sigfooter(sh, rr + 6, span=8)
for k, v in {"A": 26, "B": 16, "C": 14, "D": 14, "E": 14,
             "F": 16, "G": 14, "H": 14}.items():
    sh.column_dimensions[k].width = v

# ---- mark the eight essential inputs on the Setup sheet (column C: unused,
# ---- so no formula reference shifts)
ESSENTIAL_ROWS = [12, 19, 26, 34, 35, 37, 53, 58]
for _r in ESSENTIAL_ROWS:
    s.cell(row=_r, column=3, value="◀ START HERE")
    style(s, f"C{_r}:C{_r}", bold=True, size=9, color=WHITE, fill=GOLD, align="center")
s.column_dimensions["C"].width = 16

# ---- tab colours: green = you type here, navy = read the answer, grey = reference
for _n, _c in (("0. Start Here", GOLD), ("1. Setup", GREEN), ("2. Dashboard", NAVY),
               ("3. Scenarios", NAVY2), ("4. Trend", NAVY2),
               ("Ref", BORDER), ("Signature", SLATE)):
    wb[_n].sheet_properties.tabColor = _c

wb.active = 0

# print areas — tight, landscape, fit-to-page, so PDF/print/screenshot export
# is clean instead of dumping the full default A1:XFD grid
from openpyxl.worksheet.page import PageMargins
PRINT_AREAS = {
    "0. Start Here": "A1:H48",
    "1. Setup": "A1:H70", "2. Dashboard": "A1:H70",
    "3. Scenarios": "A1:F35", "4. Trend": "A1:M40",
    "Ref": "A1:D36", "Signature": "A1:B14",
}
for _name, _rng in PRINT_AREAS.items():
    _ws = wb[_name]
    _ws.print_area = _rng
    _ws.page_setup.orientation = "landscape"
    _ws.page_setup.fitToWidth = 1
    _ws.page_setup.fitToHeight = 1
    _ws.sheet_properties.pageSetUpPr.fitToPage = True
    _ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2,
                                   header=0, footer=0)

# doc properties
wb.properties.creator = AUTHOR
wb.properties.lastModifiedBy = AUTHOR
wb.properties.title = "CFO Premium Dashboard — Entrepreneur Business Intelligence"
wb.properties.subject = f"Digital Signature {SIG}"
wb.properties.description = (f"Created by {AUTHOR} ({GH}). Signature {SIG}. "
                             f"MIT License — attribution required.")
wb.properties.keywords = f"{AUTHOR}, {SIG}, CFO, dashboard, BDT, multi-currency"
wb.properties.category = "Financial Analysis"
wb.properties.identifier = SIG

out = ARGS.output
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
