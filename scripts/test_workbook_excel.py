#!/usr/bin/env python3
"""
test_workbook_excel.py — validate the Excel workbook, not just the calculator.

WHY THIS EXISTS
  openpyxl will happily write a workbook that Excel refuses to open, and it
  cannot evaluate a single formula. Both failures have actually happened here:

    1. A CSS-style font stack ("Noto Sans,FreeSans,Arial Unicode MS,Arial") in
       a font name made every rebuild unopenable in Excel. Excel font names
       must be one family, at most 31 characters, with no commas.
    2. Conditional-format fills were written with fgColor. Excel renders
       differential-format fills from bgColor, so the RED/GREEN status pills
       were white text on no fill - invisible from the day they shipped.

  Neither was caught by the Python tests, because neither is a Python problem.

TWO LAYERS
  WorkbookStructure - openpyxl only. Runs anywhere, including Linux CI, and
                      catches both bugs above plus sheet/formula drift.
  ExcelEngine       - drives real Excel through COM, recalculates every
                      formula, and checks Excel's own answers against
                      cfo_calc.py on identical inputs. Skips cleanly when
                      Excel or pywin32 is unavailable.

    python3 scripts/test_workbook_excel.py

Author: Md Kamrul Hasan
GitHub: https://github.com/Kamrul5242
License: MIT
Signature: MKH-EBIC-2.2.0
"""

import os
import shutil
import sys
import tempfile
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
# CFO_WORKBOOK lets CI point these tests at a fresh rebuild rather than the
# committed binary - that is the case that actually regressed.
WORKBOOK = os.environ.get(
    "CFO_WORKBOOK", os.path.join(ROOT, "assets", "cfo-premium-dashboard.xlsx"))

sys.path.insert(0, HERE)

EXPECTED_SHEETS = ["0. Start Here", "1. Setup", "2. Dashboard",
                   "3. Scenarios", "4. Trend", "Ref", "Signature"]
EXPECTED_FORMULAS = 161

# Worked example 1 from references/06-worked-examples.md, keyed by the label
# in column A of "1. Setup" so the test survives rows moving.
EXAMPLE_INPUTS = {
    "Gross Revenue": 920000,
    "Returns & Refunds": 70000,
    "Discounts Given": 0,
    "Product / Material (LANDED)": 442000,
    "Direct Labour": 0,
    "Packaging": 25500,
    "Outbound Shipping": 102000,
    "Payment Gateway Fees": 21250,
    "Marketplace Commission": 0,
    "RTO / Failed Delivery": 0,
    "Advertising Spend": 240000,
    "Salaries (non-production)": 120000,
    "Owner Compensation": 0,
    "Rent": 60000,
    "Software & Subscriptions": 0,
    "Other Operating": 0,
    "Depreciation & Amortization": 0,
    "Interest Expense": 0,
    "Tax": 0,
    "Number of Orders": 1700,
}


def _openpyxl():
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        return None


class WorkbookStructure(unittest.TestCase):
    """Static checks. No Excel required."""

    @classmethod
    def setUpClass(cls):
        load_workbook = _openpyxl()
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not installed")
        if not os.path.exists(WORKBOOK):
            raise unittest.SkipTest("workbook not found: %s" % WORKBOOK)
        cls.wb = load_workbook(WORKBOOK)

    def test_expected_sheets(self):
        self.assertEqual(self.wb.sheetnames, EXPECTED_SHEETS)

    def test_formula_count_has_not_drifted(self):
        n = sum(1 for ws in self.wb.worksheets
                for row in ws.iter_rows()
                for c in row
                if isinstance(c.value, str) and c.value.startswith("="))
        self.assertEqual(n, EXPECTED_FORMULAS)

    def test_font_names_are_a_single_family(self):
        """The bug that made every rebuild unopenable in Excel.

        A font name may not be a CSS-style stack. Excel's limit is 31
        characters and a comma is illegal.
        """
        bad = []
        for font in self.wb._fonts:
            name = getattr(font, "name", None)
            if not name:
                continue
            if "," in name or len(name) > 31:
                bad.append(name)
        self.assertEqual(bad, [], "illegal Excel font name(s): %r" % (bad,))

    def test_conditional_format_fills_use_bgcolor(self):
        """Differential formats render from bgColor, not fgColor.

        openpyxl reports dxf.fill as None on read, so the object API cannot see
        this; the raw styles part has to be inspected. A dxf patternFill with
        only fgColor renders as no highlight at all, which is how the RED/GREEN
        status pills silently did nothing from the day they were added.
        """
        import re
        import zipfile

        with zipfile.ZipFile(WORKBOOK) as z:
            styles = z.read("xl/styles.xml").decode("utf-8")
        block = re.search(r"<dxfs\b.*?</dxfs>", styles, re.S)
        if not block:
            self.skipTest("workbook defines no differential styles")
        fills = re.findall(r"<patternFill\b.*?</patternFill>|<patternFill\b[^>]*/>",
                           block.group(0), re.S)
        self.assertTrue(fills, "conditional formatting defines no fills")
        offenders = [f for f in fills if "bgColor" not in f]
        self.assertEqual(
            offenders, [],
            "conditional-format fill(s) omit bgColor and will render as "
            "nothing in Excel: %r" % (offenders,))

    def test_signature_present_in_document_properties(self):
        blob = " ".join(str(x) for x in (
            self.wb.properties.creator, self.wb.properties.subject,
            self.wb.properties.description, self.wb.properties.keywords,
            self.wb.properties.identifier))
        self.assertIn("Md Kamrul Hasan", blob)
        self.assertIn("MKH-EBIC", blob)


def _excel():
    """Return a fresh hidden Excel instance, or None if unavailable."""
    try:
        import win32com.client
    except ImportError:
        return None
    try:
        app = win32com.client.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        return app
    except Exception:
        return None


class ExcelEngine(unittest.TestCase):
    """Checks that need Excel's own calculation engine."""

    @classmethod
    def setUpClass(cls):
        if os.name != "nt":
            raise unittest.SkipTest("Excel automation is Windows-only")
        if not os.path.exists(WORKBOOK):
            raise unittest.SkipTest("workbook not found")
        cls.app = _excel()
        if cls.app is None:
            raise unittest.SkipTest("Excel or pywin32 unavailable")
        # Work on a copy. The shipped workbook is signed; never mutate it.
        cls.tmp = os.path.join(tempfile.mkdtemp(), "wb-under-test.xlsx")
        shutil.copyfile(WORKBOOK, cls.tmp)
        try:
            cls.wb = cls.app.Workbooks.Open(cls.tmp)
        except Exception as exc:
            cls.app.Quit()
            raise AssertionError(
                "Excel refused to open the workbook: %s\n"
                "This is the failure mode an illegal font name produces." % exc)
        cls.app.CalculateFullRebuild()

    @classmethod
    def tearDownClass(cls):
        try:
            cls.wb.Close(False)
        except Exception:
            pass
        try:
            cls.app.Quit()
        except Exception:
            pass
        shutil.rmtree(os.path.dirname(cls.tmp), ignore_errors=True)

    # ---- helpers ----

    def _row_of(self, sheet, label, column=1):
        ws = self.wb.Worksheets(sheet)
        for r in range(1, ws.UsedRange.Rows.Count + 1):
            if str(ws.Cells(r, column).Value) == label:
                return r
        self.fail("label %r not found in column %d of %r" % (label, column, sheet))

    def _set_example_inputs(self):
        ws = self.wb.Worksheets("1. Setup")
        for label, value in EXAMPLE_INPUTS.items():
            ws.Cells(self._row_of("1. Setup", label), 2).Value = value
        self.app.CalculateFullRebuild()

    # ---- tests ----

    def test_opens_with_expected_sheets(self):
        names = [self.wb.Worksheets(i + 1).Name
                 for i in range(self.wb.Worksheets.Count)]
        self.assertEqual(names, EXPECTED_SHEETS)

    def test_no_cell_evaluates_to_an_error(self):
        errors = []
        for i in range(self.wb.Worksheets.Count):
            ws = self.wb.Worksheets(i + 1)
            for cell in ws.UsedRange:
                text = cell.Text
                if isinstance(text, str) and text.startswith("#"):
                    errors.append("%s!%s = %s" % (ws.Name, cell.Address(0, 0), text))
        self.assertEqual(errors, [], "formula errors: %r" % (errors[:10],))

    def test_excel_agrees_with_cfo_calc(self):
        """The cross-engine check: two independent implementations, one answer."""
        import cfo_calc
        self._set_example_inputs()

        args = cfo_calc.build_parser().parse_args([
            "margins", "--revenue", "920000", "--returns", "70000",
            "--cogs", "467500", "--variable", "123250",
            "--adspend", "240000", "--opex", "180000",
        ])
        expected = cfo_calc.margins(args)

        def dash(label):
            return float(self.wb.Worksheets("2. Dashboard")
                         .Cells(self._row_of("2. Dashboard", label), 2).Value)

        self.assertAlmostEqual(dash("Net Revenue"), expected["net_revenue"], 2)
        self.assertAlmostEqual(dash("Gross Profit"), expected["gross_profit"], 2)
        self.assertAlmostEqual(dash("Contribution (pre-ad)"),
                               expected["contribution_after_variable"], 2)
        self.assertAlmostEqual(
            dash("CM after Ads"),
            expected["contribution_after_variable"] - expected["ad_spend"], 2)
        self.assertAlmostEqual(dash("Operating Profit"),
                               expected["operating_profit_ebit"], 2)
        self.assertAlmostEqual(dash("Net Profit"), expected["net_profit"], 2)
        # The worked example's headline result.
        self.assertAlmostEqual(dash("Operating Profit"), -160750.0, 2)

    def test_trend_sheet_computes_the_worked_example(self):
        ws = self.wb.Worksheets("4. Trend")
        ws.Range("B5").Value = 850000    # net revenue
        ws.Range("B6").Value = 467500    # cogs
        ws.Range("B7").Value = 123250    # variable
        ws.Range("B8").Value = 240000    # ad spend
        ws.Range("B9").Value = 180000    # fixed opex
        self.app.CalculateFullRebuild()
        self.assertAlmostEqual(float(ws.Range("B11").Value), 259250.0, 2)
        self.assertAlmostEqual(float(ws.Range("B13").Value), -160750.0, 2)
        self.assertAlmostEqual(float(ws.Range("B15").Value), -160750.0, 2)

    def test_status_pills_actually_render(self):
        """A status cell must show text AND be visibly highlighted.

        Text alone is not enough: the pills were bold white on no fill, so they
        read as blank to a human while Cells().Text still returned "LOSS".
        DisplayFormat reflects conditional formatting as applied on screen.
        """
        self._set_example_inputs()
        ws = self.wb.Worksheets("2. Dashboard")
        r = self._row_of("2. Dashboard", "Operating Profit")
        cell = ws.Cells(r, 3)
        self.assertIn(str(cell.Text).strip(), ("OK", "LOSS", "FLAT"))

        font_colour = cell.DisplayFormat.Font.Color
        fill_colour = cell.DisplayFormat.Interior.Color
        # xlNone comes back as a very large value; white is 16777215.
        self.assertNotEqual(
            fill_colour, 16777215,
            "status pill has a white background, so its white text is invisible")
        self.assertNotEqual(
            font_colour, fill_colour,
            "status pill text and background are the same colour")


if __name__ == "__main__":
    unittest.main(verbosity=2)
